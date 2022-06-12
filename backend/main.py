from ast import Return
from datetime import datetime
import imp
from io import BytesIO
from typing import List, Union
from openpyxl import load_workbook
import json
from fastapi import FastAPI, HTTPException, UploadFile, File
from bd import pelicula
from bd.empleado import fetch_login, get_empleado_by_name, get_or_create_cargo, get_or_create_dependencia, get_or_create_empleado, get_or_create_eps, get_or_create_pension, insertar_novedades, get_empleados
from handlers.handle_libro import handle_libros
from models.empleado import Cargo, Dependencia, Empleado, Eps, Pension, Novedad
import bd.book as bdb
from models.libro import Autor, Idioma, Libro, Publicador
from models.pelicula import Pelicula, Genero
from utils.empleado_parser import tuple_to_cargo, tuple_to_dependencia, tuple_to_empleado, tuple_to_eps, tuple_to_pension
from utils.generos_parser import tuple_to_genero
from dataclasses import asdict
from utils.sqlconnector import mydb
from fastapi.middleware.cors import CORSMiddleware


app = FastAPI()

origins = [
    "*",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.post("/bookFile")
async def read_book_doc(libros: UploadFile):
    json_data = json.load(libros.file)
    handle_libros(json_data, mydb)

    return {'data': len(json_data)}

@app.post("/userFile")
async def read_employees_doc(documento_empleados: bytes = File()):
    wb = load_workbook(BytesIO(documento_empleados))
    for sheet in wb:
        if sheet.title == 'MAESTRO NÓMINA':
            for row in sheet.iter_rows(min_row=3, values_only=True):
                ## Creation of employee and dependencies

                dep = tuple_to_dependencia(get_or_create_dependencia(asdict(Dependencia(nombre=row[2])), mydb))
                cargo = tuple_to_cargo(get_or_create_cargo(asdict(Cargo(nombre=row[3])), mydb))
                eps = tuple_to_eps(get_or_create_eps(asdict(Eps(nombre=row[5])), mydb))
                pension = tuple_to_pension(get_or_create_pension(asdict(Pension(nombre=row[7])), mydb))
                
                get_or_create_empleado(asdict(Empleado(row[0], row[8], row[6], datetime.strptime(str(row[4]), '%Y%m%d'), row[1],"admin", 2, cargo, pension, dep, eps)),mydb)
        else:
            lista_novedades: List[Novedad] = []
            for row in sheet.iter_rows(min_row=3, values_only=True):
                codigo_empleado = row[0]
                cod_novedad = 3
                if row[2] != None:
                    cod_novedad = 2
                elif row[3] != None:
                    cod_novedad = 1

                dias_trabajados = row[3]
                dias_no_trabajados = 0
                if row[4] != None:
                    dias_no_trabajados = row[4]
                else: 
                    dias_no_trabajados = row[5]
                
                # fecha inicio
                fecha_inicio= None
                if (row[6] != None): fecha_inicio = datetime.strptime(row[6], '%Y.%m.%d')
                elif (row[8] != None): fecha_inicio = datetime.strptime(row[8], '%Y.%m.%d')
                else: fecha_inicio = None

                # fecha fin
                fecha_fin= None
                if (row[6] != None): fecha_fin = datetime.strptime(row[7], '%Y.%m.%d')
                elif (row[8] != None): fecha_fin = datetime.strptime(row[9], '%Y.%m.%d')
                else: fecha_fin = None

                bonificacion = float(row[10])
                transporte = float(row[11])

                novedad = Novedad(
                    cod_tipo_novedad=cod_novedad,
                    dias_trabajados=dias_trabajados,
                    dias_no_trabajados= dias_no_trabajados,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,
                    bonificacion=float(bonificacion),
                    transporte=float(transporte),
                    cod_empleado=codigo_empleado,
                )
                lista_novedades.append(novedad)
            insertar_novedades([asdict(i) for i in lista_novedades], mydb)

@app.post("/movieFile")
async def read_movies_dec(documento_peliculas:bytes= File()):
    list_doc = documento_peliculas.decode().split("\n")
    list_generos: List[Genero] = []
    list_peliculas: List[Pelicula] = []
    for i in list_doc:
        datos = i.split("::")
        if len(datos) <= 1: continue
        n_pelicula = Pelicula(codigo=int(datos[0]),nombre=datos[1])
        for gen in datos[2].split("|"):
            new_gen = Genero(nombre=gen)
            n_pelicula.add_genero(new_gen)
            if (not new_gen in list_generos): list_generos.append(new_gen)
        list_peliculas.append(n_pelicula)
    pelicula.create_generos([asdict(i) for i in list_generos], mydb)
    gen_upd_gen = [tuple_to_genero(i) for i in pelicula.get_generos(mydb)]
    for i in list_peliculas:
        i.update_generos(gen_upd_gen)
    pelicula.add_peliculas([asdict(i) for i in list_peliculas], mydb)
    return {'data': len(list_peliculas)}
    


@app.get("/empleado")
async def retrieve_empleados(pagination:int = -1, quantity:int = -1, fecha_inicio:str = "", fecha_fin:str = "", name:str = ""):
    empleados:List[Empleado] = [tuple_to_empleado(i, mydb) for i in get_empleados(quantity, pagination, fecha_inicio, fecha_fin, name, mydb)]
    for i in empleados:
        i.fecha_ingreso = i.fecha_ingreso.strftime("%Y-%m-%d")
    return empleados

@app.get("/empleado/{nombre}")
async def get_empleado(nombre:str = ""):
    empleados:List[Empleado] = [tuple_to_empleado(i, mydb) for i in get_empleado_by_name(nombre, mydb)]
    for i in empleados:
        i.fecha_ingreso = i.fecha_ingreso.strftime("%Y-%m-%d")
    return empleados

@app.get("/peliculas")
async def get_peliculas(pagination:int=-1, quantity:int=-1, title:str=""):
    datos = pelicula.get_peliculas(pagination, quantity, title, mydb)
    print(datos)
    lista_pelis:List[Pelicula] = []
    if len(datos) > 1:
        new_peli = Pelicula(datos[0][0], datos[0][1], [Genero(codigo=datos[0][3], nombre=datos[0][2])])
        for i in datos[1:]:
            if new_peli.codigo != i[0]:
                lista_pelis.append(new_peli)
                new_peli = Pelicula(i[0], i[1], [Genero(codigo=i[3], nombre=i[2])])
            else:
                new_peli.add_genero(Genero(codigo=i[3], nombre=i[2]))
        lista_pelis.append(new_peli)
    elif len(datos) == 1: return [Pelicula(datos[0][0], datos[0][1], [Genero(codigo=datos[0][3], nombre=datos[0][2])])]
    return lista_pelis

@app.get("/libros")
async def get_libros(pagination:int=-1, quinatity:int=-1, title:str="", fecha_inicio:str="", fecha_fin:str=""):
    tuple_data = bdb.get_libros(pagination, quinatity, title, fecha_inicio, fecha_fin, mydb)
    libros:List[Libro] = []
    new_libro=Libro()
    for i in tuple_data:
        new = Libro(
            book_id=i[0],
            num_pages=i[1],
            average_rating=i[2],
            title=i[3],
            isbn=i[4],
            isbn_13=i[5],
            publication_date=i[6],
            ratings_count=i[7],
            text_reviews_count=i[8],
            idioma=Idioma(codigo=i[11], nombre=i[12]),
            publicador=Publicador(codigo=i[13], nombre=i[14]),
        )
        if new_libro.book_id == new.book_id:
            new.add_autor(Autor(codigo=i[9], nombre=i[10]))
            libros.append(new)
            new_libro = new
        else:
            new_libro.add_autor(Autor(codigo=i[9], nombre=i[10]))
            new_libro = new
    for i in libros:
        if (type(i.publication_date) != str):
            i.publication_date = i.publication_date.strftime("%Y-%m-%d")
    return libros

@app.post("/login")
async def login(codigo:str, password:str):
    tuple_info = fetch_login(codigo, password, mydb)
    if len(tuple_info) == 0:
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    user = tuple_to_empleado(tuple_info[0], mydb)
    user.fecha_ingreso = user.fecha_ingreso.strftime("%Y-%m-%d")
    return user